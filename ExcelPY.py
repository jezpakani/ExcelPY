from datetime import datetime, date, time, timedelta
from dateutil.parser import parse
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from os import system, name
from colorama import init, Fore
from random import randint

import argparse


class ExcelPY:
    def __init__(self):
        """
        Initialize our class and get things ready.
        """
        self.start_time = datetime.now()  # start timer before first operation
        self.end_time = datetime.now()  # end timer after last operation
        self.execution_time = datetime.now()  # results of our timer

        # filenames for our workbooks (spreadsheets)
        self.fn_alm = 'dump-alm.xlsx'
        self.fn_defect = 'dump-defects.xlsx'
        self.fn_enhancement = 'dump-enhancements.xlsx'
        self.fn_incident = 'dump-incidents.xlsx'
        self.fn_destination = 'capacity-tracker.xlsx'

        # workbook handles
        self.wb_alm = Workbook()
        self.wb_defect = Workbook()
        self.wb_enhancement = Workbook()
        self.wb_incident = Workbook()
        self.wb_destination = Workbook()

        # user supplied options (arguments)
        self.arg_data = False  # generate test data only
        self.arg_check = False  # allows us to run the app to test files without writing to them

        # general application options
        self.date_fields = ['opened', 'planned fix date']
        self.test_data_row_count = 3  # how many rows of test data we will be creating
        self.cells_updated = 0  # number of cells that were updated
        self.rows_appended = 0  # number of rows that were appended
        self.errors = 0  # how many errors were encountered
        self.warnings = 0  # counter for how many warnings were generated

    def __del__(self):
        """
        Perform cleanup operations when we destroy our class instance.
        :return: None
        """
        self.close_files()

    def open_workbooks(self):
        """
        Open files needed to perform our processes.
        :return: boolean
        """
        try:
            self.wb_alm = load_workbook(self.fn_alm)
            self.wb_defect = load_workbook(self.fn_defect)
            self.wb_enhancement = load_workbook(self.fn_enhancement)
            self.wb_incident = load_workbook(self.fn_incident)
            self.wb_destination = load_workbook(self.fn_destination)

            self.wb_alm.iso_dates = True
            self.wb_defect.iso_dates = True
            self.wb_enhancement.iso_dates = True
            self.wb_incident.iso_dates = True
            self.wb_destination.iso_dates = True
        except Exception as e:
            self.error(str(e))
            return False

        return True

    def close_files(self):
        """
        Close all of the files that we used.
        :return: None
        """
        self.wb_alm.close()
        self.wb_defect.close()
        self.wb_enhancement.close()
        self.wb_incident.close()
        self.wb_destination.close()

    def generate_test_data(self):
        """
        Populate our input file with random test data. Due to the different number of columns in the files
        and the random nature of this method there will likely be zero matching data between the files.
        :return: None
        """
        self.message('Generating {} rows of unique keyed test data.'.format(self.test_data_row_count))
        if not self.open_workbooks():
            exit()

        # populate our data dump input files
        self.populate_sheet(self.wb_incident, self.wb_incident.active, self.fn_incident, 'Hypercare Incidents', 'INC')
        self.populate_sheet(self.wb_enhancement, self.wb_enhancement.active, self.fn_enhancement,
                            'Hypercare Enhancements', 'ENH')
        self.populate_sheet(self.wb_defect, self.wb_defect.active, self.fn_defect, 'Hypercare Defects', 'DFC')
        self.populate_sheet(self.wb_alm, self.wb_alm.active, self.fn_alm, 'ALM Defects', 'ALM')

        self.message('Completed generating input file')

    def populate_sheet(self, wb, ws, fn, tab, extra):
        """
        Create test data.
        :param wb: the workbook to be used
        :param ws: the worksheet to be used
        :param fn: the file name to save changes to
        :param tab: name of the destination tab (sheet) to write to
        :param extra: extra info to add to the generated data
        :return: None
        """
        ws.delete_rows(2, ws.max_row + 1)
        self.wb_destination.active = self.wb_destination[tab]
        self.wb_destination.active.delete_rows(2, self.wb_destination.active.max_row + 1)
        used = []

        try:
            # this loop is for the dump file
            for x in range(2, self.test_data_row_count + 2):  # now lets generate some cell data
                for y in range(1, ws.max_column + 1):
                    column_header = ws.cell(row=1, column=y).value
                    if column_header is None:  # no idea why but some sheets not reporting column count correctly.
                        break

                    rand_x = randint(100, 99999)
                    rand_y = randint(100, 99999)
                    buffer = '[{}] {}:{}'.format(extra, str(rand_x).zfill(5), str(rand_y).zfill(5))

                    while used.count(buffer) != 0:  # disallow duplicate keys
                        rand_x = randint(100, 99999)
                        rand_y = randint(100, 99999)
                        buffer = '[{}] {}:{}'.format(extra, str(rand_x).zfill(5), str(rand_y).zfill(5))

                    used.append(buffer)  # add our new key to the list

                    if column_header.lower() in self.date_fields:
                        buffer = date.today() + timedelta(days=randint(-10, 35))
                        buffer = self.format_date(buffer.strftime('%Y-%m-%d'))

                    ws.cell(row=x, column=y).value = buffer

                    if y == 1:  # only write to destination the first time through so keys match
                        self.wb_destination.active.cell(row=x, column=1).value = buffer

                # this loop is for the destination file
                for y in range(2, self.wb_destination.active.max_column + 1):
                    column_header = self.wb_destination.active.cell(row=1, column=y).value
                    if column_header is None:  # no idea why but some sheets not reporting column count correctly.
                        break

                    rand_x = randint(100, 99999)
                    rand_y = randint(100, 99999)
                    buffer = '[{}] {}:{}'.format(extra, str(rand_x).zfill(5), str(rand_y).zfill(5))

                    while used.count(buffer) != 0:  # disallow duplicate keys
                        rand_x = randint(100, 99999)
                        rand_y = randint(100, 99999)
                        buffer = '[{}] {}:{}'.format(extra, str(rand_x).zfill(5), str(rand_y).zfill(5))

                    used.append(buffer)  # add our new key to the list

                    if column_header.lower() in self.date_fields:
                        buffer = date.today() + timedelta(days=randint(-10, 35))
                        buffer = self.format_date(buffer.strftime('%Y-%m-%d'))
                    else:
                        buffer = '[{}] {}:{}'.format(extra, str(rand_x).zfill(5), str(rand_y).zfill(5))

                    self.wb_destination.active.cell(row=x, column=y).value = buffer

            # now set the column widths
            for x in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(x)].width = 18

            for x in range(1, self.wb_destination.active.max_column + 1):
                self.wb_destination.active.column_dimensions[get_column_letter(x)].width = 18

        except IndexError as e:
            self.error(str(e))
        except Exception as e:
            self.error(str(e))

        wb.save(fn)
        self.wb_destination.active = self.wb_destination['Hypercare Incidents']
        self.wb_destination.save(self.fn_destination)

    def get_execution_time(self):
        """
        Display the results of our execution timer.
        :return: None
        """
        self.execution_time = self.end_time - self.start_time

        print('\n')
        self.message('**[OPERATION COMPLETE]**********************************************************************')
        if self.arg_data:
            self.message(' Execution Time: {} ms'.format(self.execution_time))
            self.message('********************************************************************************************')
        else:
            self.message('   Cell Updates: {}'.format(self.cells_updated))
            self.message(' Cell Additions: {}'.format(self.rows_appended))
            self.message('         Errors: {}'.format(self.errors))
            self.message('       Warnings: {}'.format(self.warnings))
            self.message(' Execution Time: {} ms'.format(self.execution_time))
            self.message('********************************************************************************************')

    def start_timer(self):
        """
        Start our timer used to determine execution speed.
        :return: None
        """
        self.start_time = datetime.now()

    def stop_timer(self):
        """
        Stop our timer used to determine execution speed.
        :return:
        """
        self.end_time = datetime.now()

    def is_not_used(self):
        """
        Useful to get rid of the warning messages about 'self' not being used in a method.
        :return:
        """
        pass

    def parse_args(self):
        """
        By default without args we will process the files, but we will also give options via
        command line arguments for things like generating test data.
        :return: None
        """
        parser = argparse.ArgumentParser()
        parser.add_argument('-d', '--data', dest='data',
                            help='Generate requested amount of test data.',
                            type=int, nargs='+')
        parser.add_argument('-c', '--check', action='store_true',
                            dest='check', help='Check files without modifying them.',
                            default=False)
        args = parser.parse_args()
        self.arg_data = args.data
        self.arg_check = args.check

        if xc.arg_data:  # did the user request to generate test data?
            choice = input(Fore.YELLOW + 'This option will ' + Fore.RED +
                           '*OVERWRITE ALL FILES* ' + Fore.YELLOW + 'you sure (y/n)? ')
            if choice.upper() == 'Y':
                self.test_data_row_count = int(self.arg_data[0])
                xc.generate_test_data()
            else:
                xc.arg_data = False
        else:
            self.process_dump_files()

    def process_dump_files(self):
        """
        Process all the sheets in our workbooks looking for changes.
        :return: None
        """
        if not self.open_workbooks():
            exit()

        self.message('*****************************************************************************')
        self.message('Only columns that exist in both the dump and destination file will be synced.')
        self.message('They must also match exactly including spelling and capitalization.')
        self.message('*****************************************************************************')

        self.wb_destination.active = self.wb_destination['Hypercare Incidents']
        self.parse_dump_file(self.wb_incident.active, self.wb_destination.active, self.fn_incident)

        self.wb_destination.active = self.wb_destination['Hypercare Defects']
        self.parse_dump_file(self.wb_defect.active, self.wb_destination.active, self.fn_defect)

        self.wb_destination.active = self.wb_destination['Hypercare Enhancements']
        self.parse_dump_file(self.wb_enhancement.active, self.wb_destination.active, self.fn_enhancement)

        self.wb_destination.active = self.wb_destination['ALM Defects']
        self.parse_dump_file(self.wb_alm.active, self.wb_destination.active, self.fn_alm)

    def parse_dump_file(self, ws_dump, ws_dest, fn_dump):
        """
        The actual meat of the application this method performs the synchronization of our files.
        :param ws_dump: one of our dump workbooks that we will use as data input
        :param ws_dest: the name of the worksheet in our output file that should be parsed
        :param fn_dump: the name of the dump file being parsed
        :return: None
        """
        self.message('BEGIN: [{}] -> [{}]:'.format(fn_dump.upper(), self.fn_destination.upper()), True)
        dump_headers = {}  # column headers from our dump file
        dest_headers = {}  # column headers from our destination file
        comm_headers = {}  # column headers common to both files

        rows_updated = 0
        rows_appended = 0

        # get a list of dump column headers so we can use them for searching
        for x, cell in enumerate(ws_dump[1]):
            dump_headers[cell.value] = x + 1

        # get a list of destination column headers so we can use them for searching
        for x, cell in enumerate(ws_dest[1]):
            dest_headers[cell.value] = x + 1

        # get a list of column headers from both sheets using column locations from destination
        for key1, val1 in dest_headers.items():
            for key2, val2 in dump_headers.items():
                if key1 == key2:
                    comm_headers[key1] = val1
                    break

        # now parse our dump file to arg_check for duplicate 'keys'
        if self.worksheet_has_duplicate_keys(ws_dump, fn_dump):
            return

        # now parse our destination file to arg_check for duplicate 'keys'
        if self.worksheet_has_duplicate_keys(ws_dest, self.fn_destination):
            return

        # let's case-sensitive arg_check our column headers for differences if any.
        s1 = set(dump_headers)
        s2 = set(dest_headers)

        if s1 != s2:
            s1_diff = (s1 - s2)
            s2_diff = (s2 - s1)
            if len(s1_diff) > 0:
                self.warning('{} exclusively contains the following columns: '.format(fn_dump.upper()))
                for x, item in enumerate(s1_diff):
                    self.warning('\t{}. \'{}\''.format(x + 1, str(item)))
            if len(s2_diff) > 0:
                self.warning('{} exclusively contains the following columns: '.format(self.fn_destination.upper()))
                for x, item in enumerate(s2_diff):
                    self.warning('\t{}. \'{}\''.format(x + 1, str(item)))

        dump_dict = self.parse_worksheet_into_dictionary(ws_dump, comm_headers)
        dest_dict = self.parse_worksheet_into_dictionary(ws_dest, comm_headers)
        comb_dict = {**dest_dict, **dump_dict}

        for a, b in enumerate(comb_dict):  # enumerate dictionary key
            key = b
            dump_row = dump_dict[key]
            for c, d in enumerate(dump_row):  # enumerate dictionary rows
                value = dump_row[d]['value']
                result = self.get_cell_details(ws_dest, key, d)
                if result['cell_found']:  # does this key exist in destination
                    this = ws_dest.cell(row=dest_dict[b][d]['row'], column=dest_dict[b][d]['col'])
                    if this.value != value:  # update destination cell
                        self.cells_updated += 1
                        self.format_cell_updated(this, value)
                    else:
                        self.format_cell_reset(this)
                else:  # key is not present so we are creating a new row
                    if result['key_found']:  # we need to add the remaining values for columns
                        this = ws_dest.cell(row=result['row'], column=result['col'])
                    else:  # append a new row and add the primary key
                        this = ws_dest.cell(row=ws_dest.max_row + 1, column=comm_headers[d])

                    self.format_cell_updated(this, value)
                    self.rows_appended += 1

                if self.is_date(this.value):
                    cell_date = self.format_date(this.value)
                    today = self.format_date(datetime.now())
                    if cell_date < today:
                        self.format_cell_date_passed(this)

        # save our workbook with all changes
        self.cells_updated += rows_updated
        self.rows_appended += rows_appended
        self.message('END: [{}]  [Updates: {}] [Additions: {}]'.format(fn_dump.upper(), rows_updated, rows_appended))

        # set the active worksheet so it opens on this tab
        if not self.arg_check:
            self.wb_destination.active = self.wb_destination['Hypercare Incidents']
            self.wb_destination.save(self.fn_destination)

    def get_cell_details(self, ws, primary_key, header_name):
        """
        This method does not match values, it only retrieves the failure found at the intersection
        of primary_key,header name which is effectively row,col.
        :param ws: the worksheet to search
        :param primary_key: the key field to search
        :param header_name: the column header to search
        :return: ['key_found'] - we matched primary key
                 ['grid_found'] - we matched primary key and column header
                 ['cell_found'] - we used row, col grid and that cell's value was not None
        """
        results = {'cell_found': False, 'key_found': False, 'grid_found': False}

        try:
            for x in range(2, ws.max_row + 1):
                if primary_key == ws.cell(row=x, column=1).value:  # true if we find our primary key
                    results['key_found'] = True
                    for y in range(1, ws.max_column + 1):
                        if header_name == ws.cell(row=1, column=y).value:  # true if we found our column header
                            results['grid_found'] = True
                            results['row'] = x
                            results['col'] = y
                            if ws.cell(row=x, column=y).value is not None:  # true if there is a value at grid location
                                results['value'] = ws.cell(row=x, column=y).value
                                results['cell_found'] = True
                            return results
        except Exception as e:
            self.error(str(e))

        return results

    def parse_worksheet_into_dictionary(self, ws, headers):
        """
        To avoid using loops all over the place we load our worksheets into dictionaries and run all logic from
        there. This gives us the useful information in our return value.
        :param ws: worksheet to parse
        :param headers: list of common headers in order to match destination columns
        :return: ['value'] - the value in each cell
                 ['row'] - the row for each cell
                 ['col'] - the column for each cell
        """
        self.is_not_used()
        result = {}

        try:
            for x in range(2, ws.max_row + 1):  # enumerate rows
                data = {}
                pkey = ws.cell(row=x, column=1).value
                for key, value in headers.items():
                    buffer = self.get_cell_details(ws, pkey, key)
                    if buffer['cell_found']:
                        value = {'value': buffer['value'], 'row': buffer['row'], 'col': buffer['col']}
                        data[key] = value
                        result[pkey] = data

        except Exception as e:
            self.error(str(e))

        return result

    def worksheet_has_duplicate_keys(self, ws, fn):
        """
        Parse a worksheet (primarily useful for dump files) and check for duplicate primary keys.
        :param ws: the works sheet to parse
        :param fn: the fle name associated with ws
        :return: boolean
        """
        self.is_not_used()
        results = {}

        for x in ws.iter_rows(2, ws.max_row, values_only=True):  # enumerate our worksheet keys
            key = x[0]
            if key in results:  # see if key is already in the dictionary
                results[key] = results[key] + 1  # if yes then increment found counter
            else:
                results[key] = 1  # key wasn't in the dictionary so add it now

        for key, value in list(results.items()):  # enumerate our keys
            if results[key] == 1:  # if value > 1 then it is a duplicate key
                del results[key]  # not a duplicate so remove from dictionary
            else:
                results[key] = 'occurrences: ' + str(value)

        if len(results.keys()) > 0:
            self.error(
                '[{}] ({}) contains the following duplicate keys in the first column:'.format(fn.upper(), ws.title))
            self.error(str(results))
            return True
        else:
            return False

    def format_cell_updated(self, cell, value=None):
        """
        Format a cell to identify it as updated
        :param cell: the cell to format
        :param value: an optional value to set the cell to
        :return: None
        """
        self.is_not_used()
        if value is not None:
            cell.value = value

        cell.fill = PatternFill(start_color='7fffd4', end_color='7fffd4', fill_type='solid')
        cell.font = Font(name='Ubuntu', size=11, color='555555', bold=False, italic=False)

    def format_cell_reset(self, cell, value=None):
        """
        Reset a cell format to nothing
        :param cell: the cell to format
        :param value: an optional value to set the cell to
        :return: None
        """
        self.is_not_used()
        if value is not None:
            cell.value = value

        cell.fill = PatternFill(fill_type='none')
        cell.font = Font(name='Ubuntu', size=11, color='2e2e2e', bold=False, italic=False)

    def format_cell_date_passed(self, cell, value=None):
        """
        Reset a cell to indicate the date is before today
        :param cell: the cell to format
        :param value: an optional value to set the cell to
        :return: None
        """
        self.is_not_used()
        if value is not None:
            cell.value = value

        cell.fill = PatternFill(start_color='b22222', end_color='b22222', fill_type='solid')
        cell.font = Font(name='Ubuntu', size=11, color='ffffff', bold=False, italic=False)

    def days_between(self, d1, d2):
        """
        Calculate the number of days between two dates
        :param d1: the first date
        :param d2: the second date
        :return: number of days between dates
        """
        self.is_not_used()
        try:
            d1 = self.format_date(d1)
            d2 = self.format_date(d2)
            d1 = datetime.strptime(d1, '%Y-%m-%d')
            d2 = datetime.strptime(d2, '%Y-%m-%d')
            return abs((d2 - d1).days)
        except Exception as e:
            self.error(str(e))

    def format_date(self, date_val):
        """
        Generally keep our dates in the format that we want.
        :param date_val: the date value to format
        :return: a newly formatted date value
        """
        try:
            if type(date_val) is not datetime:
                d = date.fromisoformat(date_val[0:10])
            else:
                d = date_val
            return d.strftime('%Y-%m-%d')
        except Exception as e:
            self.error((str(e)))

    def is_date(self, string, fuzzy=False):
        """
        Allows us to programmatically determine if a string is a date
        :param string: the string to evaluate if a date or not
        :param fuzzy: determines if the logic should include fuzzy
        :return: boolean
        """
        self.is_not_used()
        try:
            parse(string, fuzzy=fuzzy)
            return True

        except ValueError:
            return False

    def message(self, value='', line_before=False):
        """
        Format general messages including attributes.
        :param value: the string to display as a message
        :param line_before: boolean if a blank line should appear before the message
        :return: None
        """
        self.is_not_used()
        if line_before:
            print('\n')
        print(Fore.GREEN + '+++ ' + value)

    def error(self, value='', line_before=False):
        """
        Format general errors including attributes.
        :param value: the string to display as a message
        :param line_before: boolean if a blank line should appear before the message
        :return: None
        """
        self.errors += 1
        if line_before:
            print('\n')
        print(Fore.RED + '!!! ' + value)

    def warning(self, value='', line_before=False):
        """
        Format general warnings including attributes.
        :param value: the string to display as a message
        :param line_before: boolean if a blank line should appear before the message
        :return: None
        """
        self.warnings += 1
        if line_before:
            print('\n')
        print(Fore.YELLOW + '--- ' + value)


def clear_screen():
    """
    Clear the screen taking into account operating system.
    :return: None
    """
    if name == "nt":
        system('cls')
    else:
        system('clear')


if __name__ == '__main__':
    clear_screen()
    init(autoreset=True)
    xc = ExcelPY()
    xc.start_timer()
    xc.parse_args()
    xc.stop_timer()
    xc.get_execution_time()
    del xc
    exit(0)
